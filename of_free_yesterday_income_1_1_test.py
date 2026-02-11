from datetime import date, timedelta, datetime
import sys
import os
import time
import pyautogui
import pandas as pd
import subprocess
import atexit
import openpyxl
import yfinance as yf
from pathlib import Path
from playwright.sync_api import sync_playwright
from playwright.sync_api import Page


# region Script to help build the executable with PyInstaller
try:
    # For PyInstaller executable
    sys.path.append(os.path.join(sys._MEIPASS, "repository"))
except Exception:
    # For development
    sys.path.append(str(Path(__file__).resolve().parent.parent / "repository"))

# Global variables to keep references
playwright_instance = None
browser_context = None
# endregion

def cleanup_playwright():
    """Function to clean up Playwright resources on exit"""
    global playwright_instance, browser_context
    try:
        if browser_context:
            browser_context.close()
        if playwright_instance:
            playwright_instance.stop()
    except:
        pass

# Register cleanup function to be called on exit
atexit.register(cleanup_playwright)

def open_chrome_with_profile():
    """
    Opens a Google Chrome instance with all profile data,
    keeping the session active indefinitely.
    """
    global playwright_instance, browser_context

    # Install Playwright if necessary
    #print("Checking Playwright installation...")
    subprocess.run(["playwright", "install"], shell=True, capture_output=True)
    subprocess.run(["playwright", "install", "chromium"], shell=True, capture_output=True)

    # Chrome profile path
    profile_path = r"C:\Users\danie\AppData\Local\Google\Chrome\User Data\Default"

    # Check if the profile directory exists
    if not os.path.exists(profile_path):
        raise FileNotFoundError(f"Profile directory not found: {profile_path}")

    print("Warning: Make sure Chrome is completely closed")

    try:
        #print("Starting Playwright...")
        # IMPORTANT: Do NOT use 'with' here to keep the session active
        playwright_instance = sync_playwright().start()

        #print("Launching Chrome with profile...")

        # Launch Chrome with specific profile
        browser_context = playwright_instance.chromium.launch_persistent_context(
            user_data_dir=profile_path,
            headless=False,
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            args=[
                '--no-first-run',
                '--no-default-browser-check',
                '--disable-background-timer-throttling',
                '--disable-backgrounding-occluded-windows',
                '--disable-renderer-backgrounding',
                '--disable-sync',
                '--disable-web-security',
                '--disable-features=msRealtimeCommunication,TranslateUI',
                '--remote-debugging-port=0'
            ],
            timeout=30000
        )

        #print("Browser launched successfully!")

        # Open new tab
        page = browser_context.new_page()
        #print("New tab opened")

        # Navigate to the page
        #print("Navigating to OnlyFans messages page...")
        page.goto("https://onlyfans.com", timeout=15000)

        #print("Chrome opened successfully with all profile data!")
        #print("You can now interact with cookies, history, and saved profile data")

        return browser_context, page

    except Exception as e:
        print(f"Error opening Chrome: {e}")
        print("Troubleshooting tips:")
        print("1. Make sure Chrome is completely closed (check Task Manager)")
        print("2. Run this script as administrator")
        print("3. Try using a different profile path if necessary")

        # Clean up resources in case of error
        cleanup_playwright()
        raise

def keep_browser_alive():
    """
    Keeps the browser active indefinitely.
    Call this function after open_chrome_with_profile() if you want to keep the browser open.
    """
    try:
        print("Keeping browser active... (Press Ctrl+C to terminate)")
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nClosing browser...")
        cleanup_playwright()

def insert_username(page):
    """
    Attempt to find the username input field and insert 'hacksimone29@gmail.com'.
    Updated to target standard Google login inputs (specifically #input-13).
    """
    try:
        # List of selectors to try (updated with new ID #input-13 and standard attributes)
        selectors = [
            # 1. Direct ID match (Most specific based on your update)
            "#input-13",
            "input#input-13",

            # 2. Standard Google Email Input attributes (High reliability)
            "input[type='email'][name='email']",
            "input[name='email']",
            "input[type='email']",

            # 3. XPath specific to the new ID
            "//*[@id='input-13']",

            # 4. XPath general for email inputs
            "//input[@type='email' and @name='email']",
            "//input[@name='email']",

            # 5. Legacy/Shadow DOM selectors (Kept just in case the structure reverts or varies)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'email\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-jnygnm9")'
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            // Silent fail for this selector
                        }}
                        return false;
                    }}''', "hacksimone29@gmail.com")
                    if input_inserted:
                        print("✓ Username inserted successfully with JS/Shadow selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility just in case
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click() # Click ensures focus
                            xpath_elements.first.fill("hacksimone29@gmail.com")
                            print(f"✓ Username inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click() # Click ensures focus
                            css_elements.first.fill("hacksimone29@gmail.com")
                            print(f"✓ Username inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                # Continue to next selector if one fails
                continue

        # Fallback JavaScript approach with comprehensive search (Updated for #input-13)
        print("Trying JavaScript fallback approach for username input...")
        fallback_inserted = page.evaluate('''(text) => {
            // 1. Try Direct ID first (Fastest)
            const directInput = document.getElementById('input-13');
            if (directInput) {
                directInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                directInput.focus();
                directInput.value = text;
                directInput.dispatchEvent(new Event('input', { bubbles: true }));
                directInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 2. Try Standard Name/Type attributes
            const standardInput = document.querySelector('input[name="email"]') || document.querySelector('input[type="email"]');
            if (standardInput) {
                standardInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                standardInput.focus();
                standardInput.value = text;
                standardInput.dispatchEvent(new Event('input', { bubbles: true }));
                standardInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 3. Try Shadow DOM (Legacy support)
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                const shadowInput = shadowHost.shadowRoot.querySelector('input[type="email"]');
                if (shadowInput) {
                    shadowInput.value = text;
                    shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                    return true;
                }
            }

            return false;
        }''', "hacksimone29@gmail.com")

        if fallback_inserted:
            print("✓ Username inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into username input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_username: {str(e)}")
        return False

def insert_password(page):
    """
    Attempt to find the password input field and insert 'Booegabi10!'.
    Updated to target standard Google login inputs (specifically #input-16).
    """
    # The new defined password
    PASSWORD_TEXT = "Booegabi10!"

    try:
        # List of selectors to try (updated with new ID #input-16 and standard attributes)
        selectors = [
            # 1. Direct ID match (Most specific based on your update)
            "#input-16",
            "input#input-16",

            # 2. Standard Password Input attributes (High reliability)
            "input[name='password']",
            "input[type='password']",

            # 3. Handle case where type might be 'text' (as seen in your snippet) but name is password
            "input[type='text'][name='password']",

            # 4. XPath specific to the new ID
            "//*[@id='input-16']",

            # 5. XPath general for password inputs
            "//input[@name='password']",
            "//input[@type='password']",

            # 6. Legacy/Shadow DOM selectors (Kept just in case)
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("input[type=\'password\']")',
            'document.querySelector("#privacy-web-auth").shadowRoot.querySelector("#floating-input-sekcpj1")'
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Handle different selector types
                if selector.startswith("document.querySelector"):
                    # JavaScript selector (handles shadow DOM)
                    input_inserted = page.evaluate(f'''(text) => {{
                        try {{
                            const input = {selector};
                            if (input) {{
                                input.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                                input.focus();
                                input.value = text;
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                                return true;
                            }}
                        }} catch(e) {{
                            // Silent fail
                        }}
                        return false;
                    }}''', PASSWORD_TEXT)
                    if input_inserted:
                        print("✓ Password inserted successfully with JS/Shadow selector")
                        return True

                elif selector.startswith('/'):
                    # XPath selector
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.evaluate(
                                    `{selector}`,
                                    document,
                                    null,
                                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                                    null
                                ).singleNodeValue;
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            xpath_elements.first.scroll_into_view_if_needed()
                            xpath_elements.first.click()
                            xpath_elements.first.fill(PASSWORD_TEXT)
                            print(f"✓ Password inserted successfully with XPath: {selector}")
                            return True
                        except Exception as e:
                            print(f"XPath insert failed: {str(e)}")

                else:
                    # CSS selector
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        try:
                            # Force visibility
                            page.evaluate(f'''(selector) => {{
                                const element = document.querySelector(selector);
                                if (element) {{
                                    element.style.opacity = '1';
                                    element.style.visibility = 'visible';
                                    element.style.display = 'block';
                                }}
                            }}''', selector)

                            css_elements.first.scroll_into_view_if_needed()
                            css_elements.first.click()
                            css_elements.first.fill(PASSWORD_TEXT)
                            print(f"✓ Password inserted successfully with CSS selector: {selector}")
                            return True
                        except Exception as e:
                            print(f"CSS selector insert failed: {str(e)}")

            except Exception as e:
                continue

        # Fallback JavaScript approach with comprehensive search (Updated for #input-16)
        print("Trying JavaScript fallback approach for password input...")
        fallback_inserted = page.evaluate('''(text) => {
            // 1. Try Direct ID first
            const directInput = document.getElementById('input-16');
            if (directInput) {
                directInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                directInput.focus();
                directInput.value = text;
                directInput.dispatchEvent(new Event('input', { bubbles: true }));
                directInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 2. Try Standard Name/Type attributes
            // Note: We check name="password" first as it's safer than type="text"
            const standardInput = document.querySelector('input[name="password"]') ||
                                  document.querySelector('input[type="password"]');

            if (standardInput) {
                standardInput.scrollIntoView({behavior: 'smooth', block: 'center'});
                standardInput.focus();
                standardInput.value = text;
                standardInput.dispatchEvent(new Event('input', { bubbles: true }));
                standardInput.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }

            // 3. Try Shadow DOM (Legacy support)
            const shadowHost = document.querySelector("#privacy-web-auth");
            if (shadowHost && shadowHost.shadowRoot) {
                const shadowInput = shadowHost.shadowRoot.querySelector('input[type="password"]');
                if (shadowInput) {
                    shadowInput.value = text;
                    shadowInput.dispatchEvent(new Event('input', { bubbles: true }));
                    return true;
                }
            }

            return false;
        }''', PASSWORD_TEXT)

        if fallback_inserted:
            print("✓ Password inserted successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or insert into password input using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in insert_password: {str(e)}")
        return False

def click_on_login_button(page):
    """
    Attempt to find and click the 'Log in' button.
    Strategies: Custom attributes, text content, and standard submit types.
    """
    try:
        # List of selectors to try
        selectors = [
            # 1. Custom attribute (Very specific and likely stable)
            "button[at-attr='submit']",

            # 2. Text content (Playwright specific, very robust)
            "button:has-text('Log in')",

            # 3. Standard Submit Button
            "button[type='submit']",

            # 4. CSS Classes (Combination of classes)
            "button.g-btn.m-rounded.m-block",

            # 5. XPath by text
            "//button[contains(text(), 'Log in')]",

            # 6. XPath by type
            "//button[@type='submit']"
        ]

        # Try each selector
        for selector in selectors:
            try:
                # Check if element exists and is visible
                if page.locator(selector).count() > 0:
                    btn = page.locator(selector).first
                    if btn.is_visible():
                        print(f"✓ Found login button with selector: {selector}")

                        # Scroll and Click
                        btn.scroll_into_view_if_needed()

                        # Optional: Force wait for button to be enabled
                        if btn.is_enabled():
                            btn.click()
                            print("✓ Clicked login button successfully.")
                            return True
                        else:
                            print(f"⚠ Button found ({selector}) but is disabled.")
            except Exception as e:
                # Ignore minor errors during search
                continue

        # Fallback JavaScript approach
        print("Trying JavaScript fallback approach for login button...")
        fallback_clicked = page.evaluate('''() => {
            // 1. Try by custom attribute
            const btnAttr = document.querySelector("button[at-attr='submit']");
            if (btnAttr) {
                btnAttr.click();
                return true;
            }

            // 2. Try by text content
            const buttons = Array.from(document.querySelectorAll('button'));
            const loginBtn = buttons.find(b => b.innerText.includes('Log in'));
            if (loginBtn) {
                loginBtn.click();
                return true;
            }

            // 3. Try by type submit
            const btnSubmit = document.querySelector("button[type='submit']");
            if (btnSubmit) {
                btnSubmit.click();
                return true;
            }

            return false;
        }''')

        if fallback_clicked:
            print("✓ Login button clicked successfully using JavaScript fallback!")
            return True

        print("❌ Could not find or click the login button using any method.")
        return False

    except Exception as e:
        print(f"❌ Error in click_on_login_button: {str(e)}")
        return False


def sum_yesterday_earnings(page: Page) -> tuple[float, str]:
    """
    Soma os ganhos líquidos (Net) das entradas da data de ontem na tabela.
    Tenta formato em inglês e português.

    :param page: Objeto Page do Playwright com a página carregada.
    :return: Tuple com (soma total dos ganhos de ontem, formato de data usado)
    """
    try:
        from datetime import datetime, timedelta

        # Prepare yesterday's date in both formats
        yesterday = datetime.now() - timedelta(days=1)

        # English format: "Feb 10, 2026"
        yesterday_en = yesterday.strftime("%b %d, %Y")

        # Portuguese format: "10 fev, 2026"
        months_pt = {
            1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
            7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
        }
        yesterday_pt = f"{yesterday.day} {months_pt[yesterday.month]}, {yesterday.year}"

        print(f"\nSearching for earnings from yesterday...")
        print(f"  English format: {yesterday_en}")
        print(f"  Portuguese format: {yesterday_pt}")

        # Wait for table to load
        page.wait_for_selector("table.b-table.m-responsive.m-earnings", timeout=10000)
        page.wait_for_timeout(2000)

        # Extract all rows data using JavaScript
        rows_data = page.evaluate('''() => {
            const rows = document.querySelectorAll('table.b-table.m-responsive.m-earnings tbody tr');
            const results = [];

            rows.forEach(row => {
                // Skip the infinite loading row
                if (row.querySelector('td[colspan]')) return;

                // Get date element
                const dateSpan = row.querySelector('td.b-table__date span.b-table__date__date span');
                // Get net value element
                const netSpan = row.querySelector('td.b-table__net strong span');

                if (dateSpan && netSpan) {
                    // Clean up whitespace
                    const dateText = dateSpan.textContent.replace(/\s+/g, ' ').trim();
                    const netText = netSpan.textContent.replace(/\s+/g, ' ').trim();

                    results.push({
                        date: dateText,
                        net: netText
                    });
                }
            });

            return results;
        }''')

        if not rows_data:
            print("✗ No rows found in the table")
            return 0.0, "none"

        print(f"✓ Found {len(rows_data)} total rows in table")

        # Try to match with both date formats
        yesterday_earnings = 0.0
        yesterday_count = 0
        date_format_used = None

        for row in rows_data:
            table_date = row['date']
            net_text = row['net']

            # Check if this row matches yesterday in English OR Portuguese
            is_yesterday = (table_date == yesterday_en or table_date == yesterday_pt)

            if is_yesterday:
                # Determine which format was found
                if date_format_used is None:
                    date_format_used = "English" if table_date == yesterday_en else "Portuguese"

                yesterday_count += 1
                try:
                    # Remove $, commas, and whitespace, then convert to float
                    net_value = float(net_text.replace('$', '').replace(',', '').strip())
                    yesterday_earnings += net_value
                    print(f"  ✓ Found earning #{yesterday_count}: ${net_value:.2f} (date: {table_date})")
                except ValueError as e:
                    print(f"  ✗ Error parsing net value '{net_text}': {e}")

        if yesterday_count > 0:
            date_used = yesterday_en if date_format_used == "English" else yesterday_pt
            print(f"\n✓ SUCCESS!")
            print(f"  Date format: {date_format_used}")
            print(f"  Total transactions: {yesterday_count}")
            print(f"  Total earnings: ${yesterday_earnings:.2f}")
            return yesterday_earnings, date_used
        else:
            print(f"\n✗ No transactions found for yesterday")
            print(f"  Tried: {yesterday_en} (English) and {yesterday_pt} (Portuguese)")
            # Show first 3 dates found for debugging
            if len(rows_data) > 0:
                print(f"\n  First 3 dates in table:")
                for i, row in enumerate(rows_data[:3]):
                    print(f"    Row {i+1}: '{row['date']}'")
            return 0.0, "none"

    except Exception as e:
        print(f"✗ Error in sum_yesterday_earnings: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0.0, "none"


def get_dollar_yesterday():
    # Ticker for USD to BRL in Yahoo Finance
    ticker_symbol = "BRL=X"

    # Calculate dates
    today = datetime.now()
    yesterday = today - timedelta(days=1)

    # We request a slightly larger window (3 days) to ensure we get the latest available close
    # in case "yesterday" was a weekend or holiday.
    start_date = today - timedelta(days=5) 

    try:
        # Download data
        data = yf.download(ticker_symbol, start=start_date, end=today, progress=False)

        if data.empty:
            return None

        # Get the last available closing price (which represents the most recent "yesterday" data)
        last_quote = data['Close'].iloc[-1]

        # Extract the float value safely
        if hasattr(last_quote, 'item'):
            rate = last_quote.item()
        else:
            rate = float(last_quote)

        return round(rate, 4)

    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

def update_report(value_to_be_inserted):
    """
    Opens the Excel file, finds the next empty row in Column D, 
    inserts the value, and saves the file.
    """
    file_path = r"G:\Meu Drive\Financeiro\receita_bruta_diaria.xlsx"

    try:
        # Check if file exists to avoid crashing
        if not os.path.exists(file_path):
            print(f"Error: The file was not found at: {file_path}")
            return False

        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Gets the active sheet

        # Find the next empty row specifically in column E (index 5)
        column_e = 5
        row = 1

        # Loop until we find a cell that is None (empty)
        while ws.cell(row=row, column=column_e).value is not None:
            row += 1

        # Insert the value
        ws.cell(row=row, column=column_e, value=value_to_be_inserted)

        # Save the workbook
        wb.save(file_path)
        print(f"Successfully inserted R$ {value_to_be_inserted:.4f} into row {row}, column E.")
        return True

    except PermissionError:
        print("Error: Permission denied. Please close the Excel file and try again.")
        return False
    except Exception as e:
        print(f"Error in update_report: {str(e)}")
        return False


def main():
    # region Close any Chrome Browser instances
    time.sleep(2)
    print("Closing browser instances...")
    os.system("taskkill /f /im chrome.exe")
    time.sleep(2)
    # endregion

    browser_context, page = open_chrome_with_profile()

    # Keep browser active - IMPORTANT!
    print("Browser started successfully!")
    print("IMPORTANT: Do not close this terminal to keep Chrome active")

    # If you want the script to run indefinitely:
    # keep_browser_alive()

    # Maximize browser window
    pyautogui.press("f11")

    time.sleep(5)

    # region Login operations (if necessary)

    # region Try to insert username with retries
    print("\nAttempting to insert username...")
    max_retries = 3
    username_inserted = False

    for attempt in range(max_retries):
        print(f"Username attempt {attempt + 1}/{max_retries}")

        if insert_username(page):
            username_inserted = True
            print("✓ Username insertion confirmed.")
            break
        else:
            print(f"✗ Username attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting 2 seconds before next attempt...")
                time.sleep(2)

    if not username_inserted:
        print("❌ Failed to insert username after all attempts. Maybe you are already logged in or the selector changed again.")

    # Short pause to ensure UI processes the input before clicking "Next"
    time.sleep(2)
    # endregion

    # region Try to insert password with retries
    print("\nAttempting to insert password...")
    max_retries = 3
    password_inserted = False

    # Wait a bit for Google's transition animation (from user to password)
    time.sleep(2)

    for attempt in range(max_retries):
        print(f"Password attempt {attempt + 1}/{max_retries}")

        if insert_password(page):
            password_inserted = True
            print("✓ Password insertion confirmed.")
            break
        else:
            print(f"✗ Password attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting 2 seconds before next attempt...")
                time.sleep(2)

    if not password_inserted:
        print("❌ Failed to insert password after all attempts.")

    # Pause to ensure the field is filled before trying to click the final login button
    time.sleep(2)
    # endregion

    # region Try to click login button with retries
    print("\nAttempting to click login button...")
    max_retries = 3
    login_clicked = False

    for attempt in range(max_retries):
        print(f"Login button attempt {attempt + 1}/{max_retries}")

        if click_on_login_button(page):
            login_clicked = True
            break
        else:
            print(f"✗ Login button attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                print("Waiting 2 seconds before next attempt...")
                time.sleep(2)

    if login_clicked:
        print("✓ Login process initiated. Waiting for navigation...")
        try:
            page.wait_for_url("https://onlyfans.com/*", timeout=30000)  # Wait for post-login redirect
            page.wait_for_load_state('networkidle', timeout=30000)  # Wait for network idle to fully load
            if "login" in page.url or page.locator("text=Error").count() > 0:  # Check for error
                print("❌ Login failed: Still on login page or error detected.")
            else:
                print("✓ Login successful! Proceeding to earnings.")
        except Exception as e:
            print(f"⚠ Error in login verification: {e}")
    else:
        print("❌ Failed to click the login button.")
    time.sleep(5)  # Short pause before navigating
    # endregion

    # endregion for login process if necessary

    page.goto("https://onlyfans.com/my/statements/earnings", timeout=15000)

    # region Try to sum yesterday's earnings with retries
    max_retries = 3
    total_yesterday = 0.0
    date_format = "none"

    for attempt in range(max_retries):
        print(f"\n--- Attempt {attempt + 1}/{max_retries} ---")
        total_yesterday, date_format = sum_yesterday_earnings(page)

        if total_yesterday > 0:
            print(f"\n✓ Successfully summed yesterday's earnings on attempt {attempt + 1}")
            break
        else:
            print(f"\n✗ Attempt {attempt + 1} failed to find earnings.")
            if attempt < max_retries - 1:
                print("Waiting 2 seconds before next attempt...")
                page.wait_for_timeout(2000)
    else:
        # This else executes only if loop completes without breaking
        print("\n✗ Failed to sum yesterday's earnings after all attempts.")
        total_yesterday = 0.0

    print(f"\n{'='*50}")
    print(f"FINAL RESULT")
    print(f"{'='*50}")
    print(f"Total earnings from yesterday: ${total_yesterday:.2f}")
    if date_format != "none":
        print(f"Date format detected: {date_format}")
    print(f"{'='*50}\n")

    page.wait_for_timeout(3000)
    # endregion

    print("Fetching yesterday's dollar rate...")
    rate = get_dollar_yesterday()

    if rate:
        print(f"The dollar rate for yesterday (or last closing) was: R$ {rate}")
    else:
        print("Could not retrieve the dollar rate.")
        # You might want to exit here if rate is critical
        # sys.exit(1)

    if total_yesterday is not None and rate is not None:
        value_to_be_inserted = total_yesterday * rate
        print(f"Value to be inserted into the report: R$ {value_to_be_inserted}")

        # Here you can add the code to save to Excel if you wish
    else:
        print("Error: Could not calculate total. Missing income data or exchange rate.")

    update_report(value_to_be_inserted)

    # ADDITION: Pause the script to keep the browser open until manually closed
    print("\n✓ Automation completed. The browser will remain open until you close it manually.")
    print("The script is paused. Close the browser manually to end the process.")
    while True:
        time.sleep(1)  # Infinite loop to keep the script running without closing the browser

if __name__ == "__main__":
    main()

